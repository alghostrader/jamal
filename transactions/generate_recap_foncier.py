"""Genere les deux recapitulatifs de transaction fonciere (leasehold et freehold).

Meme structure dans les deux cas : prix vendeur / prix client avec commission,
frais de notaire (taxes gouvernementales + honoraires avec forfait plancher),
frais de geometre, total de l'operation et acompte pour l'appel de fonds.
La seule difference est la duree du bail, qui n'existe pas en freehold.
"""

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment

ARIAL = "Arial"
IDR = '#,##0'
EUR = '#,##0.00'
PCT = '0.0%'
ARES = '#,##0.00'

TAUX_CHANGE = 20788.62      # reference BCE du 20/08/2026
CHANGE_SRC = ("Taux de reference BCE du 20/08/2026 : 1 EUR = 20 788,62 IDR "
              "(source : api.frankfurter.dev).\n"
              "A actualiser au taux du jour / au taux retenu avec le client.")


def build(freehold, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Recapitulatif"

    def put(cell, value, fmt=None, bold=False):
        c = ws[cell]
        c.value = value
        c.font = Font(name=ARIAL, size=11, bold=bold)
        if fmt:
            c.number_format = fmt
        return c

    regime = "FREEHOLD (pleine propriete)" if freehold else "LEASEHOLD (bail 30 ans)"
    unite = "IDR / are" if freehold else "IDR / are / an"

    # --- En-tete ---
    put("A1", "RECAPITULATIF TRANSACTION FONCIERE - APPEL DE FONDS", bold=True)
    put("A2", "Regime : " + regime, bold=True)
    put("A3", "Client :")
    put("A4", "Terrain / localisation :")
    put("A5", "Date :")

    # --- Parametres ---
    put("A7", "PARAMETRES (cellules a modifier)", bold=True)
    params = [("surface", "Surface du terrain (ares)", 6, ARES),
              ("m2", "Surface du terrain (m2)", "=B8*100", IDR)]
    if not freehold:
        params.append(("duree", "Duree du leasehold (annees)", 30, IDR))
    params += [
        ("pv", "Prix vendeur (%s)%s" % (unite, " - A SAISIR" if freehold else ""),
         None if freehold else 4000000, IDR),
        ("pc", "Prix client avec commission (%s)%s" % (unite, " - A SAISIR" if freehold else ""),
         None if freehold else 5000000, IDR),
        ("fx", "Taux de change (IDR pour 1 EUR)", TAUX_CHANGE, '#,##0.00'),
        ("taxes", "Taxes gouvernementales (% du prix client)", 0.05, PCT),
        ("hono", "Honoraires du notaire (% du prix client)", 0.01, PCT),
        ("mini", "Honoraires du notaire - forfait minimum (IDR)", 10000000, IDR),
        ("geo", "Frais de geometre (EUR)", 1000, EUR),
        ("acompte", "Taux d'acompte (appel de fonds)", 0.10, PCT),
    ]
    P = {}
    r = 8
    for key, label, value, fmt in params:
        put("A%d" % r, label)
        put("B%d" % r, value, fmt)
        P[key] = "$B$%d" % r
        r += 1
    p_first, p_last = 8, r - 1

    ws[P["fx"].replace("$", "")].comment = Comment(CHANGE_SRC, "Agence")
    ws[P["taxes"].replace("$", "")].comment = Comment(
        "Taxes gouvernementales : 5% du prix client (prix avec commission). "
        "Hypothese indiquee par l'agent.", "Agence")
    ws[P["hono"].replace("$", "")].comment = Comment(
        "Honoraires du notaire : 1% du prix client (prix avec commission). "
        "Hypothese indiquee par l'agent.", "Agence")
    ws[P["mini"].replace("$", "")].comment = Comment(
        "Forfait plancher : si le pourcentage donne moins de 10 000 000 IDR, "
        "ce forfait s'applique a la place du pourcentage.", "Agence")
    if freehold:
        note = ("Prix d'achat en pleine propriete, exprime en IDR par are (pas de duree). "
                "A renseigner avec le prix negocie du dossier.\n"
                "Exemple : 6 ares a 900 000 000 IDR/are = 5 400 000 000 IDR.")
        ws[P["pv"].replace("$", "")].comment = Comment(note, "Agence")
        ws[P["pc"].replace("$", "")].comment = Comment(note, "Agence")

    # --- Detail du calcul ---
    r = p_last + 2
    put("A%d" % r, "DETAIL DU CALCUL", bold=True)
    r += 1
    cols = ["Poste", "Surface (ares)"]
    if not freehold:
        cols.append("Duree (ans)")
    cols += ["Prix unitaire (%s)" % unite, "Montant IDR", "Montant EUR"]
    for i, label in enumerate(cols):
        put("%s%d" % (chr(ord("A") + i), r), label, bold=True)
    C_IDR = chr(ord("A") + len(cols) - 2)   # E en leasehold, D en freehold
    C_EUR = chr(ord("A") + len(cols) - 1)   # F en leasehold, E en freehold
    C_PU = chr(ord("A") + len(cols) - 3)    # colonne prix unitaire / taux

    def ligne(label, idr_formula, row, taux=None, bold=False):
        put("A%d" % row, label, bold=bold)
        if taux:
            put("%s%d" % (C_PU, row), taux, PCT)
        put("%s%d" % (C_IDR, row), idr_formula, IDR, bold=bold)
        put("%s%d" % (C_EUR, row), "=%s%d/%s" % (C_IDR, row, P["fx"]), EUR, bold=bold)

    # prix vendeur / prix client
    r_pv = r + 1
    r_pc = r + 2
    for row, key, label in ((r_pv, "pv", "Prix du terrain - PRIX VENDEUR"),
                            (r_pc, "pc", "Prix du terrain - PRIX CLIENT (avec commission)")):
        put("A%d" % row, label)
        put("B%d" % row, "=%s" % P["surface"], ARES)
        facteurs = "B%d" % row
        if not freehold:
            put("C%d" % row, "=%s" % P["duree"], IDR)
            facteurs += "*C%d" % row
        put("%s%d" % (C_PU, row), "=%s" % P[key], IDR)
        facteurs += "*%s%d" % (C_PU, row)
        put("%s%d" % (C_IDR, row), "=" + facteurs, IDR)
        put("%s%d" % (C_EUR, row), "=%s%d/%s" % (C_IDR, row, P["fx"]), EUR)

    r_com = r_pc + 1
    ligne("dont commission agence (ecart vendeur / client)",
          "=%s%d-%s%d" % (C_IDR, r_pc, C_IDR, r_pv), r_com)

    # --- Frais annexes ---
    r = r_com + 2
    put("A%d" % r, "FRAIS ANNEXES", bold=True)
    r_taxes = r + 1
    ligne("Taxes gouvernementales (5% du prix client)",
          "=%s%d*%s" % (C_IDR, r_pc, P["taxes"]), r_taxes, taux="=%s" % P["taxes"])

    r_hono = r_taxes + 1
    ligne("Honoraires du notaire (1% du prix client, minimum 10 000 000 IDR)",
          "=MAX(%s%d*%s,%s)" % (C_IDR, r_pc, P["hono"], P["mini"]), r_hono,
          taux="=IF(%s%d=0,0,%s%d/%s%d)" % (C_IDR, r_pc, C_IDR, r_hono, C_IDR, r_pc))
    ws["%s%d" % (C_IDR, r_hono)].comment = Comment(
        "MAX entre le pourcentage d'honoraires et le forfait minimum : si le pourcentage "
        "donne moins de 10 000 000 IDR, c'est le forfait qui s'applique.\n"
        "La colonne des taux affiche le taux reellement supporte.", "Agence")

    r_not = r_hono + 1
    ligne("Sous-total frais de notaire (taxes + honoraires)",
          "=%s%d+%s%d" % (C_IDR, r_taxes, C_IDR, r_hono), r_not,
          taux="=IF(%s%d=0,0,%s%d/%s%d)" % (C_IDR, r_pc, C_IDR, r_not, C_IDR, r_pc))

    r_geo = r_not + 1
    put("A%d" % r_geo, "Frais de geometre (forfait 1 000 EUR)")
    put("%s%d" % (C_IDR, r_geo), "=%s*%s" % (P["geo"], P["fx"]), IDR)
    put("%s%d" % (C_EUR, r_geo), "=%s" % P["geo"], EUR)

    # --- Total ---
    r_tot = r_geo + 2
    ligne("TOTAL DE L'OPERATION POUR L'ACHETEUR",
          "=%s%d+%s%d+%s%d" % (C_IDR, r_pc, C_IDR, r_not, C_IDR, r_geo), r_tot, bold=True)

    # --- Appel de fonds ---
    r = r_tot + 2
    put("A%d" % r, "APPEL DE FONDS", bold=True)
    r_ac = r + 1
    ligne("Acompte 10% du total de l'operation (reservation notaire)",
          "=%s%d*%s" % (C_IDR, r_tot, P["acompte"]), r_ac,
          taux="=%s" % P["acompte"], bold=True)
    r_solde = r_ac + 1
    ligne("Solde a regler apres acompte",
          "=%s%d-%s%d" % (C_IDR, r_tot, C_IDR, r_ac), r_solde)
    r_info = r_solde + 2
    ligne("Pour information : 10% du prix du terrain seul",
          "=%s%d*%s" % (C_IDR, r_pc, P["acompte"]), r_info)

    # --- Notes ---
    r = r_info + 2
    put("A%d" % r, "NOTES / HYPOTHESES", bold=True)
    notes = ["1 are = 100 m2 -> terrain de 600 m2."]
    if freehold:
        notes.append("Achat en FREEHOLD (pleine propriete) : prix exprime en IDR par are, "
                     "sans duree. Les deux prix (lignes %d et %d des parametres) sont a saisir."
                     % (p_first + 2, p_first + 3))
        notes.append("Exemple de saisie : 6 ares a 900 000 000 IDR/are = 5 400 000 000 IDR "
                     "de prix de terrain.")
    else:
        notes.append("Achat en LEASEHOLD : prix exprime en IDR par are et par an, "
                     "multiplie par la duree du bail.")
    notes += [
        "Taux de change en %s : reference BCE du 20/08/2026 (1 EUR = 20 788,62 IDR). "
        "A actualiser avant envoi au client." % P["fx"].replace("$", ""),
        "Frais de notaire detailles : taxes gouvernementales 5%% (%s) + honoraires du notaire "
        "1%% (%s)." % (P["taxes"].replace("$", ""), P["hono"].replace("$", "")),
        "Honoraires du notaire : le pourcentage s'applique sauf s'il donne moins que le forfait "
        "minimum de 10 000 000 IDR (%s), auquel cas le forfait est retenu (formule MAX en %s%d)."
        % (P["mini"].replace("$", ""), C_IDR, r_hono),
        "Ces taux sont appliques au prix client avec commission. Remplacer %s%d par %s%d dans les "
        "formules de frais si la base retenue est le prix vendeur." % (C_IDR, r_pc, C_IDR, r_pv),
        "Frais de geometre saisis en EUR (%s) et convertis en IDR au taux %s."
        % (P["geo"].replace("$", ""), P["fx"].replace("$", "")),
        "Seules les cellules de parametres (colonne B, lignes %d a %d) sont a saisir : tout le "
        "reste du tableau est calcule par formules et se met a jour automatiquement."
        % (p_first, p_last),
    ]
    for i, n in enumerate(notes):
        put("A%d" % (r + 1 + i), n)

    widths = {"A": 62, "B": 16, C_PU: 26, C_IDR: 18, C_EUR: 15}
    if not freehold:
        widths.setdefault("C", 13)
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    print(out)


if __name__ == "__main__":
    base = "/home/user/jamal/transactions/"
    build(False, base + "Recap_transaction_fonciere_appel_de_fonds.xlsx")
    build(True, base + "Recap_transaction_fonciere_FREEHOLD_appel_de_fonds.xlsx")
