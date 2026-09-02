"""Projection de projet multi-villas — deux versions produites par le meme code.

- Version INTERNE : commission agence sur sa propre ligne, activable, sans mise
  en page particuliere, avec les notes et hypotheses en bas de page.
- Version CLIENT : commission deja incorporee dans le prix du terrain (aucune
  trace du taux de commission), mise en page soignee (bandeaux, encadrements,
  teintes), sans notes.

Structure commune : 1. cout foncier, 2. investissement (une ligne par villa),
3. revenus previsionnels (occupation et prix de nuitee propres a chaque villa).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL = "Arial"
IDR = '#,##0'
EUR = '#,##0.00'
PCT = '0.0%'
NB = '#,##0.0'
ENT = '#,##0'
IDR_U = '#,##0" IDR"'
EUR_U = '#,##0.00" €"'

TAUX_CHANGE = 20788.62   # reference BCE du 20/08/2026
N_VILLAS = 5
NCOL = "ABCDEFG"

# Palette de la version client
BLEU = "1F3864"
GRIS = "D9D9D9"
BLEU_CLAIR = "DDEBF7"
JAUNE = "FFF2CC"
THIN = Side(style="thin", color="8EA9DB")
MED = Side(style="medium", color=BLEU)


def build(client, out):
    wb = Workbook()
    ws = wb.active
    ws.title = "Projection"
    fmt_idr = IDR_U if client else IDR
    fmt_eur = EUR_U if client else EUR

    def put(cell, value, fmt=None, bold=False, fill=None, color=None, size=11,
            align=None, border=False):
        c = ws[cell]
        c.value = value
        c.font = Font(name=ARIAL, size=size, bold=bold, color=color)
        if fmt:
            c.number_format = fmt
        if client:
            if fill:
                c.fill = PatternFill("solid", fgColor=fill)
            if border:
                c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if align:
                c.alignment = Alignment(horizontal=align)
        return c

    def bandeau(row, titre):
        """Titre de section : bandeau plein largeur en version client."""
        put("A%d" % row, titre, bold=True, fill=BLEU if client else None,
            color="FFFFFF" if client else None, size=12 if client else 11)
        if client:
            for col in NCOL[1:]:
                put("%s%d" % (col, row), None, fill=BLEU)

    def param(row, label, value, fmt, commentaire=None):
        put("A%d" % row, label)
        put("B%d" % row, value, fmt, fill=JAUNE, border=True)
        if commentaire:
            ws["B%d" % row].comment = Comment(commentaire, "Agence")

    def entetes(row, libelles):
        for col, lib in zip(NCOL, libelles):
            put("%s%d" % (col, row), lib, bold=True, fill=GRIS, border=True,
                align="center" if col != "A" else None)

    def montant(row, label, formule_idr, taux=None, total=False):
        fill = BLEU_CLAIR if total else None
        put("A%d" % row, label, bold=total, fill=fill, border=True)
        put("B%d" % row, taux, PCT, bold=total, fill=fill, border=True, align="center")
        put("C%d" % row, formule_idr, fmt_idr, bold=total, fill=fill, border=True)
        put("D%d" % row, "=C%d/$B$7" % row, fmt_eur, bold=total, fill=fill, border=True)

    def liste(cellule, valeurs, commentaire):
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(valeurs), allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[cellule])
        ws[cellule].comment = Comment(commentaire, "Agence")

    # ---------------- En-tete ----------------
    put("A1", "PROJECTION DE PROJET - VILLAS", bold=True, size=16 if client else 11,
        color=BLEU if client else None)
    put("A2", "Client :")
    put("A3", "Terrain / localisation :")
    put("A4", "Date :")
    for r in (2, 3, 4):
        put("B%d" % r, None, border=True)

    # ---------------- Parametres generaux ----------------
    bandeau(6, "PARAMÈTRES GÉNÉRAUX")
    param(7, "Taux de change (IDR pour 1 EUR)", TAUX_CHANGE, '#,##0.00',
          "Taux de référence BCE du 20/08/2026 : 1 EUR = 20 788,62 IDR "
          "(source : api.frankfurter.dev). À actualiser au taux du jour.")
    param(8, "Taxes gouvernementales (% du prix du terrain)", 0.05, PCT)
    param(9, "Honoraires du notaire (% du prix du terrain)", 0.01, PCT)
    param(10, "Honoraires du notaire - forfait minimum (IDR)", 20000000, fmt_idr,
          "Forfait plancher : si le pourcentage d'honoraires donne moins que ce montant, "
          "c'est le forfait qui s'applique.")
    param(11, "Frais de géomètre (EUR)", 1000, fmt_eur)
    param(12, "Coût de création de la PT PMA (EUR)", 2000, fmt_eur)
    param(13, "Nombre de nuits commercialisables / an", 365, ENT,
          "Base annuelle appliquée à toutes les villas (365 nuits, ou moins si la villa "
          "est fermée une partie de l'année).")

    # ---------------- 1. Cout foncier ----------------
    bandeau(15, "1. COÛT FONCIER")
    r = 16
    param(r, "Option retenue : FREEHOLD ou LEASEHOLD", "LEASEHOLD", None); k_opt = r; r += 1
    param(r, "Surface du terrain (ares)", 6, '#,##0.00'); k_surf = r; r += 1
    param(r, "Surface du terrain (m²)", "=B%d*100" % k_surf, ENT); r += 1
    suffixe = "" if client else " - prix vendeur"
    param(r, "Prix FREEHOLD%s (IDR / are)" % suffixe, None, fmt_idr); k_pfh = r; r += 1
    param(r, "Prix LEASEHOLD%s (IDR / are / an)" % suffixe,
          5000000 if client else 4000000, fmt_idr); k_plh = r; r += 1
    param(r, "Durée du leasehold (années)", 30, ENT); k_duree = r; r += 1
    if not client:
        param(r, "Commission agence - à appliquer ? (OUI / NON)", "OUI", None)
        k_com_oui = r; r += 1
        param(r, "Taux de commission agence (% du prix vendeur)", 0.25, PCT,
              "Taux appliqué au prix vendeur. 25% correspond a un prix vendeur de "
              "4 000 000 IDR/are/an revendu 5 000 000 au client. À ajuster selon le dossier.")
        k_com_tx = r; r += 1
    param(r, "Frais d'emménagement (IDR)", 0, fmt_idr,
          "Montant forfaitaire à renseigner selon le dossier."); k_emm = r; r += 1
    param(r, "Cérémonie traditionnelle (IDR)", 0, fmt_idr,
          "Cérémonie traditionnelle balinaise. Montant forfaitaire a renseigner "
          "selon le dossier."); k_cer = r; r += 1

    liste("B%d" % k_opt, ["FREEHOLD", "LEASEHOLD"],
          "Choisir FREEHOLD ou LEASEHOLD dans la liste deroulante.\n"
          "FREEHOLD -> prix du terrain = surface x B%d (prix à l'are, sans durée).\n"
          "LEASEHOLD -> prix du terrain = surface x B%d x B%d (prix à l'are et par an "
          "x durée)." % (k_pfh, k_plh, k_duree))
    if not client:
        liste("B%d" % k_com_oui, ["OUI", "NON"],
              "Case à cocher : OUI ajoute la commission agence (taux de la cellule B%d) "
              "au prix vendeur, NON la retire." % k_com_tx)

    r += 1
    entetes(r, ["Poste", "Taux", "Montant IDR", "Montant EUR"]); r += 1
    prix_terrain = ('=IF($B$%d="FREEHOLD",$B$%d*$B$%d,$B$%d*$B$%d*$B$%d)'
                    % (k_opt, k_surf, k_pfh, k_surf, k_plh, k_duree))
    if client:
        montant(r, "Prix du terrain", prix_terrain); k_base = r; r += 1
    else:
        montant(r, "Prix du terrain - PRIX VENDEUR", prix_terrain); k_pv = r; r += 1
        montant(r, "Commission agence (si OUI en B%d)" % k_com_oui,
                '=IF($B$%d="OUI",C%d*$B$%d,0)' % (k_com_oui, k_pv, k_com_tx),
                taux='=IF($B$%d="OUI",$B$%d,0)' % (k_com_oui, k_com_tx)); k_com = r; r += 1
        montant(r, "PRIX DU TERRAIN AVEC COMMISSION", "=C%d+C%d" % (k_pv, k_com),
                total=True); k_base = r; r += 1
    ws["C%d" % k_base].comment = Comment(
        "FREEHOLD : surface x prix à l'are.\n"
        "LEASEHOLD : surface x prix à l'are et par an x durée.", "Agence")

    montant(r, "Taxes gouvernementales", "=C%d*$B$8" % k_base, taux="=$B$8")
    k_tax = r; r += 1
    montant(r, "Honoraires du notaire (forfait plancher en B10)",
            "=MAX(C%d*$B$9,$B$10)" % k_base, taux="=IF(C%d=0,0,C%d/C%d)" % (k_base, r, k_base))
    ws["C%d" % r].comment = Comment(
        "MAX entre le pourcentage d'honoraires (B9) et le forfait minimum (B10) : "
        "si le pourcentage donne moins que le forfait, c'est le forfait qui s'applique.",
        "Agence")
    k_hono = r; r += 1
    montant(r, "Sous-total frais de notaire (taxes + honoraires)",
            "=C%d+C%d" % (k_tax, k_hono), taux="=IF(C%d=0,0,C%d/C%d)" % (k_base, r, k_base))
    k_not = r; r += 1
    put("A%d" % r, "Frais de géomètre (forfait)", border=True)
    put("B%d" % r, None, border=True)
    put("C%d" % r, "=$B$11*$B$7", fmt_idr, border=True)
    put("D%d" % r, "=$B$11", fmt_eur, border=True)
    k_geo = r; r += 1
    montant(r, "Frais d'emménagement", "=$B$%d" % k_emm); k_lemm = r; r += 1
    montant(r, "Cérémonie traditionnelle", "=$B$%d" % k_cer); k_lcer = r; r += 1
    montant(r, "COÛT FONCIER TOTAL",
            "=C%d+C%d+C%d+C%d+C%d" % (k_base, k_not, k_geo, k_lemm, k_lcer), total=True)
    k_foncier = r; r += 2

    # ---------------- 2. Investissement ----------------
    bandeau(r, "2. INVESTISSEMENT À CONSENTIR"); r += 1
    param(r, "Création de la PT PMA - à inclure ? (OUI / NON)", "OUI", None)
    k_pt = r
    liste("B%d" % k_pt, ["OUI", "NON"],
          "Case à cocher : OUI ajoute les 2 000 EUR de creation de la PT PMA (cellule B12) "
          "a l'investissement, NON les retire.")
    r += 2

    entetes(r, ["", "Désignation de la villa", "Construction (EUR)", "Ameublement (EUR)",
                "Total villa (EUR)", "Total villa (IDR)"]); r += 1
    k_v1 = r
    for i in range(N_VILLAS):
        rr = k_v1 + i
        put("A%d" % rr, "Villa %d" % (i + 1), bold=True, border=True)
        put("B%d" % rr, "Villa %d" % (i + 1) if i == 0 else None, fill=JAUNE, border=True)
        put("C%d" % rr, None, fmt_eur, fill=JAUNE, border=True)
        put("D%d" % rr, None, fmt_eur, fill=JAUNE, border=True)
        put("E%d" % rr, "=C%d+D%d" % (rr, rr), fmt_eur, border=True)
        put("F%d" % rr, "=E%d*$B$7" % rr, fmt_idr, border=True)
    ws["B%d" % k_v1].comment = Comment(
        "Nommer chaque villa ici (villa principale, villa 2 chambres, pool house...). "
        "La désignation est reprise automatiquement dans le tableau des revenus.\n"
        "Laisser vide les villas non utilisées : elles comptent pour zéro.", "Agence")
    ws["C%d" % k_v1].comment = Comment(
        "Budget de construction en EUR. Pour un devis exprimé en IDR, "
        "saisir la formule =montant_IDR/$B$7.", "Agence")
    ws["D%d" % k_v1].comment = Comment(
        "Budget d'ameublement en EUR. Pour un devis exprimé en IDR, "
        "saisir la formule =montant_IDR/$B$7.", "Agence")
    k_vtot = k_v1 + N_VILLAS
    put("A%d" % k_vtot, "TOTAL VILLAS", bold=True, fill=BLEU_CLAIR, border=True)
    put("B%d" % k_vtot, None, fill=BLEU_CLAIR, border=True)
    for col in "CDE":
        put("%s%d" % (col, k_vtot), "=SUM(%s%d:%s%d)" % (col, k_v1, col, k_vtot - 1),
            fmt_eur, bold=True, fill=BLEU_CLAIR, border=True)
    put("F%d" % k_vtot, "=E%d*$B$7" % k_vtot, fmt_idr, bold=True, fill=BLEU_CLAIR, border=True)
    r = k_vtot + 2

    entetes(r, ["Poste", "Taux", "Montant IDR", "Montant EUR"]); r += 1
    montant(r, "Coût foncier (report du tableau 1)", "=C%d" % k_foncier); r += 1
    montant(r, "Constructions et ameublements (total villas)", "=F%d" % k_vtot); r += 1
    montant(r, "Création de la PT PMA (si OUI en B%d)" % k_pt,
            '=IF($B$%d="OUI",$B$12*$B$7,0)' % k_pt); r += 1
    montant(r, "INVESTISSEMENT TOTAL", "=C%d+C%d+C%d" % (r - 3, r - 2, r - 1), total=True)
    k_inv = r; r += 2

    # ---------------- 3. Revenus previsionnels ----------------
    bandeau(r, "3. REVENUS PRÉVISIONNELS PAR VILLA"); r += 1
    param(r, "Charges d'exploitation (% du revenu brut)", 0.30, PCT,
          "Pourcentage appliqué au revenu brut cumulé de toutes les villas "
          "(gestion, ménage, entretien, énergie...).")
    k_charges = r; r += 2

    entetes(r, ["", "Désignation de la villa", "Taux d'occupation", "Prix nuitée (EUR)",
                "Nuits occupées / an", "Revenus bruts (EUR)", "Revenus bruts (IDR)"]); r += 1
    k_r1 = r
    for i in range(N_VILLAS):
        rr = k_r1 + i
        rv = k_v1 + i
        put("A%d" % rr, "Villa %d" % (i + 1), bold=True, border=True)
        put("B%d" % rr, '=IF($B$%d="","",$B$%d)' % (rv, rv), border=True)
        put("C%d" % rr, 0.65 if i == 0 else None, PCT, fill=JAUNE, border=True, align="center")
        put("D%d" % rr, 250 if i == 0 else None, fmt_eur, fill=JAUNE, border=True)
        put("E%d" % rr, "=$B$13*C%d" % rr, NB, border=True, align="center")
        put("F%d" % rr, "=E%d*D%d" % (rr, rr), fmt_eur, border=True)
        put("G%d" % rr, "=F%d*$B$7" % rr, fmt_idr, border=True)
    ws["C%d" % k_r1].comment = Comment(
        "Taux d'occupation propre à cette villa (ex. 65% = 237 nuits sur 365).", "Agence")
    ws["D%d" % k_r1].comment = Comment(
        "Prix moyen de la nuitée de cette villa, en EUR. Laisser vide si la villa "
        "n'est pas utilisee dans ce scenario.", "Agence")
    k_rtot = k_r1 + N_VILLAS
    put("A%d" % k_rtot, "TOTAL", bold=True, fill=BLEU_CLAIR, border=True)
    for col in "BCD":
        put("%s%d" % (col, k_rtot), None, fill=BLEU_CLAIR, border=True)
    for col, f in (("E", NB), ("F", fmt_eur), ("G", fmt_idr)):
        put("%s%d" % (col, k_rtot), "=SUM(%s%d:%s%d)" % (col, k_r1, col, k_rtot - 1),
            f, bold=True, fill=BLEU_CLAIR, border=True)
    r = k_rtot + 2

    entetes(r, ["Poste", "Taux", "Montant IDR", "Montant EUR"]); r += 1
    put("A%d" % r, "REVENUS BRUTS ANNUELS", bold=True, fill=BLEU_CLAIR, border=True)
    put("B%d" % r, None, fill=BLEU_CLAIR, border=True)
    put("C%d" % r, "=G%d" % k_rtot, fmt_idr, bold=True, fill=BLEU_CLAIR, border=True)
    put("D%d" % r, "=F%d" % k_rtot, fmt_eur, bold=True, fill=BLEU_CLAIR, border=True)
    k_brut = r; r += 1
    montant(r, "Charges d'exploitation", "=C%d*$B$%d" % (k_brut, k_charges),
            taux="=$B$%d" % k_charges)
    k_ch = r; r += 1
    montant(r, "REVENUS NETS ANNUELS", "=C%d-C%d" % (k_brut, k_ch), total=True)
    k_net = r; r += 2

    bandeau(r, "INDICATEURS"); r += 1
    for label, formule, fmt in (
            ("Nombre de villas retenues dans le scénario",
             '=COUNTIF(D%d:D%d,">0")' % (k_r1, k_rtot - 1), ENT),
            ("Rendement brut (revenus bruts / investissement total)",
             "=IF(C%d=0,0,C%d/C%d)" % (k_inv, k_brut, k_inv), PCT),
            ("Rendement net (revenus nets / investissement total)",
             "=IF(C%d=0,0,C%d/C%d)" % (k_inv, k_net, k_inv), PCT),
            ("Retour sur investissement (années)",
             "=IF(C%d=0,0,C%d/C%d)" % (k_net, k_inv, k_net), NB)):
        put("A%d" % r, label, border=True)
        put("B%d" % r, formule, fmt, bold=True, fill=BLEU_CLAIR, border=True, align="center")
        r += 1

    # ---------------- Notes (version interne uniquement) ----------------
    if client:
        r += 1
        put("A%d" % r, "Document établi à titre indicatif, sans valeur contractuelle.",
            color="808080")
    else:
        r += 1
        put("A%d" % r, "NOTES / HYPOTHESES", bold=True)
        notes = [
            "MODE D'EMPLOI : renseigner le terrain (tableau 1), nommer les villas et saisir "
            "leurs budgets (tableau 2), puis leur occupation et leur prix de nuitee "
            "(tableau 3). Les villas non utilisees restent vides et comptent pour zero.",
            "Tableau 1 : le choix FREEHOLD / LEASEHOLD en B%d pilote le calcul du prix du "
            "terrain ; le prix inutilise est simplement ignore. 1 are = 100 m2." % k_opt,
            "Les prix du terrain sont des PRIX VENDEUR : la commission agence est ajoutee "
            "separement en ligne %d, uniquement si B%d = OUI, au taux de la cellule B%d."
            % (k_com, k_com_oui, k_com_tx),
            "Les frais de notaire portent sur le prix du terrain commission comprise "
            "(ligne %d) : taxes gouvernementales (B8) + honoraires (B9) avec un forfait "
            "plancher de 20 000 000 IDR (B10)." % k_base,
            "S'ajoutent au cout foncier le geometre (B11), les frais d'emmenagement (B%d) "
            "et la ceremonie traditionnelle (B%d), saisis en IDR." % (k_emm, k_cer),
            "Tableau 2 : une ligne par villa, construction et ameublement saisis en EUR ; "
            "pour un devis en IDR, saisir =montant_IDR/$B$7. La designation saisie en "
            "colonne B est reprise automatiquement dans le tableau 3.",
            "La creation de la PT PMA (2 000 EUR, cellule B12) s'ajoute a l'investissement "
            "uniquement si B%d = OUI, quel que soit le nombre de villas." % k_pt,
            "Tableau 3 : chaque villa a son propre taux d'occupation et son propre prix de "
            "nuitee. Revenus bruts = nuits commercialisables (B13) x taux d'occupation x "
            "prix de la nuitee. Les charges s'appliquent au revenu brut cumule.",
            "Taux de change en B7 : reference BCE du 20/08/2026 (1 EUR = 20 788,62 IDR). "
            "A actualiser avant presentation au client.",
            "Projection previsionnelle a but indicatif : elle n'integre ni la fiscalite "
            "indonesienne sur les revenus locatifs, ni l'amortissement, ni les frais de "
            "gestion au-dela du pourcentage de charges saisi. En leasehold, le rendement "
            "doit s'apprecier au regard de la duree residuelle du bail.",
            "VERSION INTERNE - contient la commission agence. Ne pas transmettre au client : "
            "utiliser la version 'presentation client'.",
        ]
        for i, n in enumerate(notes):
            put("A%d" % (r + 1 + i), n, bold=(i == len(notes) - 1))

    largeurs = [("A", 50), ("B", 26), ("C", 22), ("D", 22), ("E", 20), ("F", 22), ("G", 22)]
    for col, w in largeurs:
        ws.column_dimensions[col].width = w
    if client:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"

    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    print(out)


if __name__ == "__main__":
    base = "/home/user/jamal/transactions/"
    build(False, base + "Projection_de_projet.xlsx")
    build(True, base + "Projection_de_projet_CLIENT.xlsx")
