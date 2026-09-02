"""Calcul du coût foncier et appel de fonds — fichier unique à bascule.

Une liste déroulante FREEHOLD / LEASEHOLD pilote le calcul du prix du terrain,
comme dans les projections de projet. Les cellules sans objet dans l'option
retenue sont grisées automatiquement.

Document interne : la commission agence y figure sur sa propre ligne.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL = "Arial"
IDR = '#,##0'
EUR = '#,##0.00'
PCT = '0.0%'
ENT = '#,##0'

TAUX_CHANGE = 20788.62   # référence BCE du 20/08/2026

wb = Workbook()
ws = wb.active
ws.title = "Cout foncier"


def put(cell, value, fmt=None, bold=False):
    c = ws[cell]
    c.value = value
    c.font = Font(name=ARIAL, size=11, bold=bold)
    if fmt:
        c.number_format = fmt
    return c


def param(row, label, value, fmt, commentaire=None):
    put("A%d" % row, label)
    put("B%d" % row, value, fmt)
    if commentaire:
        ws["B%d" % row].comment = Comment(commentaire, "Agence")


def montant(row, label, formule_idr, taux=None, bold=False):
    put("A%d" % row, label, bold=bold)
    if taux is not None:
        put("B%d" % row, taux, PCT)
    put("C%d" % row, formule_idr, IDR, bold=bold)
    put("D%d" % row, "=C%d/$B$7" % row, EUR, bold=bold)


def liste(cellule, valeurs, commentaire):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(valeurs), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws[cellule])
    ws[cellule].comment = Comment(commentaire, "Agence")


def etat(row, actif_si, k_opt):
    """Mention d'état en colonne C + grisage de la ligne quand elle est sans objet."""
    inactif = "FREEHOLD" if actif_si == "LEASEHOLD" else "LEASEHOLD"
    put("C%d" % row,
        '=IF($B$%d="%s",IF($B$%d>0,"","A RENSEIGNER"),"sans objet en %s")'
        % (k_opt, actif_si, row, inactif))
    ws["C%d" % row].font = Font(name=ARIAL, size=10, italic=True, color="808080")
    gris = DifferentialStyle(font=Font(color="A6A6A6", italic=True),
                             fill=PatternFill(bgColor="F2F2F2"))
    ws.conditional_formatting.add(
        "A%d:C%d" % (row, row),
        Rule(type="expression", dxf=gris, formula=['$B$%d<>"%s"' % (k_opt, actif_si)]))


# --- En-tête ---
put("A1", "CALCUL DU COÛT FONCIER ET APPEL DE FONDS", bold=True)
put("A2", "Client :")
put("A3", "Terrain / localisation :")
put("A4", "Date :")

# --- Paramètres généraux ---
put("A6", "PARAMÈTRES GÉNÉRAUX (cellules à modifier)", bold=True)
param(7, "Taux de change (IDR pour 1 EUR)", TAUX_CHANGE, '#,##0.00',
      "Taux de référence BCE du 20/08/2026 : 1 EUR = 20 788,62 IDR "
      "(source : api.frankfurter.dev). À actualiser au taux du jour.")
param(8, "Taxes gouvernementales (% du prix officiel)", 0.05, PCT)
param(9, "Honoraires du notaire (% du prix officiel)", 0.01, PCT)
param(10, "Honoraires du notaire - forfait minimum (IDR)", 20000000, IDR,
      "Forfait plancher : si le pourcentage d'honoraires donne moins que ce montant, "
      "c'est le forfait qui s'applique.")
param(11, "Frais de géomètre (EUR)", 1000, EUR)
param(12, "Taux d'acompte (appel de fonds)", 0.10, PCT)

# --- 1. Terrain ---
put("A14", "1. TERRAIN", bold=True)
param(15, "Option retenue : FREEHOLD ou LEASEHOLD", "LEASEHOLD", None)
param(16, "Surface du terrain (ares)", 6, '#,##0.00')
param(17, "Surface du terrain (m²)", "=B16*100", ENT)
param(18, "Prix FREEHOLD - prix vendeur (IDR / are)", None, IDR)
param(19, "Prix LEASEHOLD - prix vendeur (IDR / are / an)", 4000000, IDR)
param(20, "Durée du leasehold (années)", 30, ENT,
      "Utilisée uniquement en LEASEHOLD. En FREEHOLD la cellule est grisée et "
      "n'entre dans aucun calcul.")
param(21, "Commission agence - à appliquer ? (OUI / NON)", "OUI", None)
param(22, "Taux de commission agence (% du prix vendeur)", 0.25, PCT,
      "Taux appliqué au prix vendeur. 25% correspond à un prix vendeur de "
      "4 000 000 IDR/are/an revendu 5 000 000 au client. À ajuster selon le dossier.")
param(23, "Frais d'emménagement (IDR)", 0, IDR,
      "Montant forfaitaire à renseigner selon le dossier.")
param(24, "Cérémonie traditionnelle (IDR)", 0, IDR,
      "Cérémonie traditionnelle balinaise. Montant forfaitaire à renseigner selon le dossier.")

etat(18, "FREEHOLD", 15)
etat(19, "LEASEHOLD", 15)
etat(20, "LEASEHOLD", 15)
liste("B15", ["FREEHOLD", "LEASEHOLD"],
      "Choisir FREEHOLD ou LEASEHOLD dans la liste déroulante.\n"
      "FREEHOLD -> prix du terrain = surface x B18 (prix à l'are, sans durée).\n"
      "LEASEHOLD -> prix du terrain = surface x B19 x B20 (prix à l'are et par an x durée).")
liste("B21", ["OUI", "NON"],
      "Case à cocher : OUI ajoute la commission agence (taux de la cellule B22) "
      "au prix vendeur, NON la retire.")

put("A26", "Poste", bold=True)
put("B26", "Taux", bold=True)
put("C26", "Montant IDR", bold=True)
put("D26", "Montant EUR", bold=True)

montant(27, "Prix du terrain - PRIX VENDEUR",
        '=IF($B$15="FREEHOLD",$B$16*$B$18,$B$16*$B$19*$B$20)')
ws["C27"].comment = Comment(
    "FREEHOLD : surface x prix à l'are (B16 x B18).\n"
    "LEASEHOLD : surface x prix à l'are et par an x durée (B16 x B19 x B20).", "Agence")
montant(28, "Commission agence (si OUI en B21)", '=IF($B$21="OUI",C27*$B$22,0)',
        taux='=IF($B$21="OUI",$B$22,0)')
montant(29, "PRIX DU TERRAIN AVEC COMMISSION", "=C27+C28", bold=True)
montant(30, "Taxes gouvernementales (sur le prix officiel du terrain)", "=C27*$B$8",
        taux="=$B$8")
montant(31, "Honoraires du notaire (sur le prix officiel, forfait plancher en B10)",
        "=MAX(C27*$B$9,$B$10)", taux="=IF(C27=0,0,C31/C27)")
ws["C31"].comment = Comment(
    "MAX entre le pourcentage d'honoraires (B9) et le forfait minimum (B10) : "
    "si le pourcentage donne moins que le forfait, c'est le forfait qui s'applique.\n"
    "La colonne Taux affiche le taux réellement supporté.", "Agence")
montant(32, "Sous-total frais de notaire (taxes + honoraires)", "=C30+C31",
        taux="=IF(C27=0,0,C32/C27)")
put("A33", "Frais de géomètre (forfait)")
put("C33", "=$B$11*$B$7", IDR)
put("D33", "=$B$11", EUR)
montant(34, "Frais d'emménagement", "=$B$23")
montant(35, "Cérémonie traditionnelle", "=$B$24")
montant(36, "TOTAL DE L'OPÉRATION POUR L'ACHETEUR", "=C29+C32+C33+C34+C35", bold=True)

# --- 2. Appel de fonds ---
put("A38", "2. APPEL DE FONDS", bold=True)
put("A39", "Poste", bold=True)
put("B39", "Taux", bold=True)
put("C39", "Montant IDR", bold=True)
put("D39", "Montant EUR", bold=True)
montant(40, "Acompte du total de l'opération (réservation notaire)", "=C36*$B$12",
        taux="=$B$12", bold=True)
montant(41, "Solde à régler après acompte", "=C36-C40")
montant(43, "Pour information : acompte sur le prix du terrain seul", "=C29*$B$12")

# --- Notes ---
put("A45", "NOTES / HYPOTHÈSES", bold=True)
notes = [
    "Le choix FREEHOLD / LEASEHOLD en B15 pilote le calcul du prix du terrain ; le prix "
    "inutilisé (B18 ou B19/B20) est grisé et ignoré. 1 are = 100 m².",
    "FREEHOLD : prix exprimé en IDR par are, sans durée. LEASEHOLD : prix en IDR par are "
    "et par an, multiplié par la durée du bail.",
    "Les prix saisis sont des PRIX VENDEUR : la commission agence est ajoutée séparément "
    "en ligne 28, uniquement si B21 = OUI, au taux de la cellule B22.",
    "Les frais de notaire portent sur le PRIX OFFICIEL du terrain (ligne 27), commission "
    "exclue : taxes gouvernementales (B8) + honoraires (B9) avec un forfait plancher de "
    "20 000 000 IDR (B10).",
    "S'ajoutent au total le géomètre (B11), les frais d'emménagement (B23) et la cérémonie "
    "traditionnelle (B24), ces deux derniers saisis en IDR.",
    "Taux de change en B7 : référence BCE du 20/08/2026 (1 EUR = 20 788,62 IDR). "
    "À actualiser avant envoi au client.",
    "Seules les cellules de paramètres (colonne B, lignes 7 à 12 et 15 à 24) sont à saisir : "
    "tout le reste du tableau est calculé par formules.",
    "DOCUMENT INTERNE - contient la commission agence. Pour une présentation client, "
    "utiliser la projection de projet version client.",
]
for i, n in enumerate(notes):
    put("A%d" % (46 + i), n, bold=(i == len(notes) - 1))

for col, w in [("A", 58), ("B", 18), ("C", 20), ("D", 18)]:
    ws.column_dimensions[col].width = w

wb.calculation.fullCalcOnLoad = True
out = "/home/user/jamal/transactions/Cout_foncier_appel_de_fonds.xlsx"
wb.save(out)
print(out)
